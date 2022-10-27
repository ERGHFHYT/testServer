from tkinter import filedialog as fd
from openpyxl import load_workbook
import classes

def excel(the_name_of_the_table,label):
    filename = fd.askopenfilename(
        title='Open a file',
        initialdir='/',
    )
    try:
        wb = load_workbook(filename)
        ws = wb.active
        if the_name_of_the_table == "teachers_table" or "practitioners_table" or "pupils_table":
            A = ws["A"]
            B = ws["B"]
            C = ws["C"]
            exel_list = []
            for a in range(len(A)):
                if A[a].value is None:
                    A[a].value = "ריק"
                if B[a].value is None:
                    B[a].value = "ריק"
                if C[a].value is None:
                    C[a].value = "ריק"
        else:
            A = ws["A"]
            B = ws["B"]
            exel_list = []
            for a in range(len(A)):
                if A[a].value is None:
                    A[a].value = "ריק"
                if B[a].value is None:
                    B[a].value = "ריק"

        def add_item(item_list):
            for exel in exel_list:
                bool_i = True
                for P in item_list:
                    if P[1] == exel[1] and P[0] == exel[0] and P[2] == exel[2]:
                        bool_i = False
                if bool_i:
                    item_list.append(exel)
            return item_list

        for i in range(len(A)):
            if the_name_of_the_table == "password_table":
                l = [str(A[i].value), str(B[i].value)]
            else:
                l = [A[i].value, str(B[i].value), C[i].value]
            exel_list.append(l)
        data_to_add = ["add_exel", exel_list, the_name_of_the_table]
        teachers = classes.Base.sever_db(data_to_add)
        teachers = add_item(teachers)
        list_data = tuple(teachers)
        return list_data
    except:
        pass
