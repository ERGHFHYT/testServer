import random
import threading
import tkinter

from client.constants import *
import customtkinter
from client.client_classes_folder.Base import Base

from tkinter import filedialog as fd, ttk
from openpyxl import load_workbook
import pynput
import speech_recognition
from client import search_from_the_teachers, update_item


class New_wind(Base):
    def __init__(self, root, windows_name, appearance_mode,
                 default_color_theme, label,
                 the_location):
        super().__init__(root, windows_name, appearance_mode,
                         default_color_theme, label, the_location)
        self.the_name_of_the_table = TEACHERS_TABLE
        self.box = None
        self.root = root
        self.windows_name = windows_name
        self.appearance_mode = appearance_mode
        self.default_color_theme = default_color_theme
        self.list_of_buttons = None
        self.label = None
        self.label_2 = None
        self.label_3 = None
        self.the_location = None
        self.tree = None
        self.mic_button = None

    def change_the_screen_following_the_table(self, name_for_the_test,
                                              the_number_of_the_button,
                                              the_name_of_the_table):
        self.label_2.configure(text=EMPTY_SPACE)
        self.label_3.configure(text=EMPTY_SPACE)
        self.box.delete(0, tkinter.END)
        self.the_name_of_the_table = the_name_of_the_table
        self.root.label_1.destroy()
        self.root.label_1 = customtkinter.CTkLabel(master=self.root.frame_right,
                                                   text=name_for_the_test,
                                                   font=(DEFAULT_FONT,
                                                         -25))
        self.root.label_1.place(relx=0.7, rely=0.03)
        is_admin = Base.execution_server(
            [GET_TABLE, SAVE_TABLE])
        if is_admin[0][0] == "yes":
            number_of_buttons = 6
        else:
            number_of_buttons = 5
        for number in range(number_of_buttons):
            if number == the_number_of_the_button:
                self.list_of_buttons[number].configure(fg_color=BACKGROUND_COLOR)
            else:
                self.list_of_buttons[number].configure(fg_color=NUM_BLUE)
        if the_name_of_the_table == PUPILS_IN_TEACHERS:
            list_data = Base.execution_server([PUPILS_IN_TEACHERS])
        else:
            list_data = Base.execution_server([GET_TABLE, the_name_of_the_table])
        self.clear_all_form_the_tree()
        self.add_items_to_the_tree(list_data)
        if the_name_of_the_table == CIRCULATIONS_TABLE:
            list_d = []
            for l in list_data:
                list_d.append(l[1])
            self.box["values"] = tuple(list_d)
            self.clear_all_form_the_tree()
            self.add_items_to_the_tree(list_d)
        elif the_name_of_the_table == PASSWORD_TABLE:
            for d in list_data:
                if d[2] == "yes":
                    list_data.remove(d)
            for d in list_data:
                d.remove(d[2])
            self.box["values"] = tuple(list_data)
            self.clear_all_form_the_tree()
            self.add_items_to_the_tree(list_data)
        else:
            self.box["values"] = tuple(list_data)
            self.clear_all_form_the_tree()
            self.add_items_to_the_tree(list_data)

    def window_search_from_the_teachers(self):
        self.root.destroy()
        search_from_the_teachers.search()

    def button_pupils(self):
        self.tree.column("# 1", anchor='n', width=120)
        self.tree.heading("# 1", text="שם")
        self.tree.column("# 2", anchor='n', width=120)
        self.tree.heading("# 2", text="שם משפחה")
        self.tree.column("# 3", anchor='n', width=120)
        self.tree.heading("# 3", text="תעודת זהות")
        self.tree.column("# 4", anchor='n', width=120)
        self.tree.heading("# 4", text="מספר טלפון")
        self.tree.column("# 5", anchor='n', width=90)
        self.tree.heading("# 5", text="מחזור")
        self.tree.column("# 6", anchor='n', width=140)
        self.tree.heading("# 6", text="תעודת זהות (מורה)")
        New_wind.change_the_screen_following_the_table(self, "תלמידים", 1,
                                                       PUPILS_TABLE)

    def button_pupils_in_teachers(self):
        self.tree.column("# 1", anchor='n', width=120)
        self.tree.heading("# 1", text="שם")
        self.tree.column("# 2", anchor='n', width=120)
        self.tree.heading("# 2", text="מספר טלפון")
        self.tree.column("# 3", anchor='n', width=120)
        self.tree.heading("# 3", text="תעודת זהות")
        self.tree.column("# 4", anchor='n', width=120)
        self.tree.heading("# 4", text="שם (מורה)")
        self.tree.column("# 5", anchor='n', width=90)
        self.tree.heading("# 5", text="מספר טלפון (מורה)")
        self.tree.column("# 6", anchor='n', width=140)
        self.tree.heading("# 6", text="תעודת זהות (מורה)")
        New_wind.change_the_screen_following_the_table(self, "תלמידים לפי מורים", 4,
                                                       PUPILS_IN_TEACHERS)

    def button_teacher(self):
        self.tree.column("# 1", anchor='n', width=120)
        self.tree.heading("# 1", text="שם")
        self.tree.column("# 2", anchor='n', width=120)
        self.tree.heading("# 2", text="שם משפחה")
        self.tree.column("# 3", anchor='n', width=120)
        self.tree.heading("# 3", text="תעודת זהות")
        self.tree.column("# 4", anchor='n', width=120)
        self.tree.heading("# 4", text="מספר טלפון")
        self.tree.column("# 5", anchor='n', width=90)
        self.tree.heading("# 5", text="")
        self.tree.column("# 6", anchor='n', width=140)
        self.tree.heading("# 6", text="")
        New_wind.change_the_screen_following_the_table(self, "מורים", 0,
                                                       TEACHERS_TABLE)

    def button_circulations(self):
        self.tree.column("# 1", anchor='n', width=120)
        self.tree.heading("# 1", text="מחזור")
        self.tree.column("# 2", anchor='n', width=120)
        self.tree.heading("# 2", text="")
        self.tree.column("# 3", anchor='n', width=120)
        self.tree.heading("# 3", text="")
        self.tree.column("# 4", anchor='n', width=120)
        self.tree.heading("# 4", text="")
        self.tree.column("# 5", anchor='n', width=90)
        self.tree.heading("# 5", text="")
        self.tree.column("# 6", anchor='n', width=140)
        self.tree.heading("# 6", text="")
        New_wind.change_the_screen_following_the_table(self, "מחזורים", 3,
                                                       "circulations_table")

    def button_password(self):
        self.tree.column("# 1", anchor='center')
        self.tree.heading("# 1", text="שם משתמש")
        self.tree.column("# 2", anchor='center')
        self.tree.heading("# 2", text="סיסמה")
        self.tree.column("# 3", anchor='center')
        self.tree.heading("# 3", text="")
        self.tree.column("# 4", anchor='center')
        self.tree.heading("# 4", text="")
        self.tree.column("# 5", anchor='center')
        self.tree.heading("# 5", text="")
        self.tree.column("# 6", anchor='center')
        self.tree.heading("# 6", text="")

        New_wind.change_the_screen_following_the_table(self, "סיסמה", 5,
                                                       "password_table")

    def button_Practitioners(self):
        self.tree.column("# 1", anchor='n', width=120)
        self.tree.heading("# 1", text="שם")
        self.tree.column("# 2", anchor='n', width=120)
        self.tree.heading("# 2", text="שם משפחה")
        self.tree.column("# 3", anchor='n', width=120)
        self.tree.heading("# 3", text="תעודת זהות")
        self.tree.column("# 4", anchor='n', width=120)
        self.tree.heading("# 4", text="מספר טלפון")
        self.tree.column("# 5", anchor='n', width=90)
        self.tree.heading("# 5", text="")
        self.tree.column("# 6", anchor='n', width=140)
        self.tree.heading("# 6", text="")
        New_wind.change_the_screen_following_the_table(self, "מתרגלים", 2,
                                                       "practitioners_table")

    def remove_item(self):
        if self.the_name_of_the_table != PUPILS_IN_TEACHERS:
            selected = self.tree.focus()
            if str(self.tree.focus()) != "":
                values = self.tree.item(selected, 'values')
                print(values)
                if self.the_name_of_the_table == CIRCULATIONS_TABLE:
                    value = values[0]
                elif self.the_name_of_the_table == PASSWORD_TABLE:
                    value = values[1]
                else:
                    value = values[2]
                print(value)
                Base.execution_server(
                    [REMOVE_ITEM_FROM_TABLE, self.the_name_of_the_table,
                     str(value)])
                list_data = Base.execution_server([GET_TABLE,
                                                         self.the_name_of_the_table])
                if self.the_name_of_the_table == CIRCULATIONS_TABLE:
                    list_d = []
                    for l in list_data:
                        list_d.append(l[1])
                        list_data = list_d
                elif self.the_name_of_the_table == PASSWORD_TABLE:
                    for d in list_data:
                        if d[2] == "yes":
                            list_data.remove(d)
                    for d in list_data:
                        d.remove(d[2])
                self.box["values"] = tuple(list_data)
                self.clear_all_form_the_tree()
                self.add_items_to_the_tree(list_data)
            else:
                self.label.configure(text="לא בחרת נתון")
        else:
            self.label.configure(text="אי אפשר למחוק נתון מטבלה זו")
        self.root.after(4000, self.clear_label)
        self.label_2.configure(text=EMPTY_SPACE)
        self.label_3.configure(text=EMPTY_SPACE)

    def update_single(self, the_execute):
        if self.the_name_of_the_table != PUPILS_IN_TEACHERS:
            if the_execute == "chek":
                # data = Base.execution_server(
                #     [GET_TABLE, self.the_name_of_the_table])
                selected = self.tree.focus()
                global_entry = self.tree.item(selected, 'values')
                print("the global_entry ", global_entry)
                if global_entry != EMPTY_SPACE:
                    # for d in data:
                    #     if not self.the_name_of_the_table == CIRCULATIONS_TABLE:
                    #         if d[1] in global_entry:
                    #             global_entry = d
                    #             break
                    #     else:
                    #         if d[0] in global_entry:
                    #             global_entry = d
                    #             break
                    self.root.destroy()
                    update_item.update(global_entry,
                                       self.the_name_of_the_table)
                else:
                    self.label.configure(text="לא בחרת איזה נתונים לעדכן")
                    self.root.after(4000, self.clear_label)

            else:
                global_entry = EMPTY_SPACE
                self.root.destroy()
                update_item.update(global_entry, self.the_name_of_the_table)
        else:
            self.label.configure(text="אי אפשר להוסיף נתון מטבלה זו")
        self.root.after(4000, self.clear_label)
        self.label_2.configure(text=EMPTY_SPACE)
        self.label_3.configure(text=EMPTY_SPACE)

    def catching_sound(self, typed):
        data = []
        list_data = Base.execution_server(
            [GET_TABLE, self.the_name_of_the_table])
        for item in list_data:
            if typed.lower() in item:
                data.append(item)
        self.clear_all_form_the_tree()
        self.add_items_to_the_tree(data)

    def excel(self):
        filename = fd.askopenfilename(
            title='Open a file',
            initialdir='/',
        )
        try:
            print("----------")
            wb = load_workbook(filename)
            ws = wb.active
            A = ws["A"]
            print("A", A)
            if self.the_name_of_the_table in [PASSWORD_TABLE, PUPILS_TABLE,
                                              TEACHERS_TABLE,
                                              PRACTITIONERS_TABLE]:
                B = ws["B"]
            if self.the_name_of_the_table == PUPILS_TABLE:
                C = ws["C"]
                D = ws["D"]
            exel_list = []
            for current_row in range(len(A)):
                if A[current_row].value is None:
                    A[current_row].value = "ריק"
                if self.the_name_of_the_table == TEACHERS_TABLE or \
                        PUPILS_TABLE or PRACTITIONERS_TABLE or PASSWORD_TABLE:
                    if B[current_row].value is None:
                        B[current_row].value = "ריק"
                if self.the_name_of_the_table == PUPILS_TABLE:
                    if C[current_row].value is None:
                        C[current_row].value = "ריק"
                    if D[current_row].value is None:
                        D[current_row].value = "ריק"

            def add_items_form_the_exel(item_list):
                for exel in exel_list:
                    bool_i = True
                    for P in item_list:
                        if P[1] == exel[1] and P[0] == exel[0] and P[2] == \
                                exel[2]:
                            bool_i = False
                    if bool_i:
                        item_list.append(exel)
                return item_list

            for current_row in range(len(A)):
                if self.the_name_of_the_table == PUPILS_TABLE:
                    row_to_add = [A[current_row].value,
                                  str(B[current_row].value),
                                  C[current_row].value, D[current_row].value]
                else:
                    row_to_add = [str(A[current_row].value),
                                  str(B[current_row].value)]
                exel_list.append(row_to_add)
            data_to_add = ["add_exel", exel_list, self.the_name_of_the_table]
            teachers = Base.execution_server(data_to_add)
            teachers = add_items_form_the_exel(teachers)
            list_data = tuple(teachers)
            self.box = list_data
            self.label.configure(text="הקובץ שנתנת התווסף בהצלחה")
        except:
            self.label.configure(text="הקובץ שנתנת הוא לא קובץ אקסל")
        self.root.after(4000, self.clear_label)

    def clear_label(self):
        self.label.configure(text=EMPTY_SPACE)

    def voice(self):
        self.mic_button.configure(fg_color=BACKGROUND_COLOR)
        recognizer = speech_recognition.Recognizer()
        with speech_recognition.Microphone() as mic:
            try:

                recognizer.adjust_for_ambient_noise(mic, duration=0.5)
                # mouse_listener = pynput.mouse.Listener(suppress=True)
                # mouse_listener.start()
                audio = recognizer.listen(mic, phrase_time_limit=10)
                if audio != '':
                    self.box.insert(0, str(recognizer.recognize_google(audio,
                                                                       language="he")))
                    self.mic_button.configure(fg_color=NUM_BLUE)
                    self.catching_sound(str(recognizer.recognize_google(audio,
                                                                        language="he")))

            except:
                self.mic_button.configure(fg_color=NUM_BLUE)
                self.label = self.create_masages(200, 130, BACKGROUND_COLOR,
                                                 "לא זיהה קול", 0.25,
                                                 0.3, -16)
                self.root.after(4000, self.clear_label)
        # mouse_listener.stop()

    def mic(self):
        x = threading.Thread(target=self.voice, args=())
        x.start()

    def clear_all_form_the_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

    def add_items_to_the_tree(self, list_of_items):
        the_number_of_the_item = 0
        for item in list_of_items:
            self.tree.insert('', 'end', text=str(the_number_of_the_item), values=item)
            the_number_of_the_item += 1

    def check(self, event):
        typed = self.box.get()
        if self.the_name_of_the_table == PUPILS_IN_TEACHERS:
            list_data = Base.execution_server([PUPILS_IN_TEACHERS])
        else:
            list_data = Base.execution_server(
                [GET_TABLE, self.the_name_of_the_table])
        if typed == '':
            data = list_data
        else:
            data = []
            for item in list_data:
                for i in item:
                    if typed.lower() in i.lower():
                        data.append(item)
        if self.the_name_of_the_table == CIRCULATIONS_TABLE:
            list_d = []
            for l in data:
                list_d.append(l[1])
            data = list_d
        elif self.the_name_of_the_table == PASSWORD_TABLE:
            for d in data:
                if d[2] == "yes":
                    data.remove(d)
            for d in data:
                d.remove(d[2])
        self.box["values"] = data
        self.clear_all_form_the_tree()
        self.add_items_to_the_tree(data)
