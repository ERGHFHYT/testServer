import json
import socket
import customtkinter
from tkinter import *
import tkinter
from tkinter import filedialog as fd
from openpyxl import load_workbook
import speech_recognition
from PIL import Image, ImageTk
import exelfile
import new_window
import update_item
import time


class Base:
    WIDTH = 900
    HEIGHT = 600
    CHEK = "chek"

    def __init__(self, root, windows_name, appearance_mode,
                 default_color_theme, label, the_location):
        self.root = root
        self.windows_name = windows_name
        self.appearance_mode = appearance_mode
        self.default_color_theme = default_color_theme
        self.label = label
        self.the_location = the_location

    def window(self):
        self.root.geometry("900x600")
        self.root.title(self.windows_name)
        self.root.resizable(True, True)
        customtkinter.set_appearance_mode(
            self.appearance_mode)
        customtkinter.set_default_color_theme(
            self.default_color_theme)

    def create_masages(self, width, height, fg_color, text, x, y, text_size):
        self.root.label = customtkinter.CTkLabel(
            master=self.the_location,
                                                width=width,
                                              height=height,
                                              fg_color=fg_color,
                                              corner_radius=0,
                                              bg_color=("gray70",
                                                        "gray25"), text=text,
                                              text_font=("Halvetica", text_size))
        self.root.label.place(relx=x, rely=y, anchor=tkinter.CENTER)
        return self.root.label
    def change_the_text(self,text):
        self.label.configure(text=text)
        self.root.after(4000, self.clear_label)

    def clear_label(self):
        self.label.configure(text="")

    @classmethod
    def sever_db(self, message):
        host = "localhost"  # as both code is running on same pc
        port = 5050  # socket server port number
        client_socket_1 = socket.socket()  # instantiate
        client_socket_1.connect((host, port))  # connect to the server
        message = json.dumps(message)
        message = message.encode('utf-8')
        client_socket_1.send(message)  # send message
        data = client_socket_1.recv(256 * 1024).decode('utf-8')
        data = json.loads(data)
        if data:
            client_socket_1.close()
        return data


class custom_window(Base):
    def __init__(self, root, windows_name, appearance_mode,
                 default_color_theme, the_name_of_the_table, color_object,
                 label, the_location):
        super().__init__(root, windows_name, appearance_mode,
                         default_color_theme, label, the_location)
        self.root = root
        self.color_object = color_object
        self.windows_name = windows_name
        self.appearance_mode = appearance_mode
        self.default_color_theme = default_color_theme
        self.the_name_of_the_table = the_name_of_the_table
        self.label = label
        self.the_location = the_location

        tool = Base(root, windows_name, appearance_mode, default_color_theme,
                    label, the_location)
        tool.window()

    def start(self, the_names_of_the_entrys, data_entry):
        def add_item():
            if entrys[1].get().isnumeric():
                if len(str(entrys[1].get())) == 10:
                    if not entrys[0].get().isnumeric():
                        e = []
                        for entry in entrys:
                            e.append(str(entry.get()))
                        data = ["add_item", self.the_name_of_the_table, e]
                        Base.sever_db(data)
            else:
                self.label.configure(text="הכנסת קלט שהוא לא מספר")


            self.root.after(4000, self.clear_label)

        def go_back():
            self.root.destroy()
            new_window.main_window()

        self.root.frame_1 = customtkinter.CTkFrame(master=self.root,
                                                   corner_radius=0,
                                               height=550, width=650)
        self.root.frame_1.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)

        label_1 = customtkinter.CTkFrame(master=self.root.frame_1, height=550,
                                         width=650)
        label_1.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)
        entrys = []
        num = 0.3
        for name in range(len(the_names_of_the_entrys)):
            entrys.append(0)
            entrys[name] = self.root.my_entry_2 = customtkinter.CTkEntry(
                master=self.root.frame_1, corner_radius=0, width=200,
                placeholder_text=the_names_of_the_entrys[name], text_font=(
                    "Halvetica", -20),
                justify='right')
            self.root.my_entry_2.place(relx=0.5, rely=num, anchor=tkinter.CENTER)
            num += 0.15
        print(len(entrys))

        self.root.button_2 = customtkinter.CTkButton(master=self.root.frame_1,
                                                 text="עדכן",
                                                corner_radius=0,
                                                     command=add_item,
                                                     width=200)
        self.root.button_2.place(relx=0.7, rely=0.85, anchor=tkinter.CENTER)
        self.root.button_3 = customtkinter.CTkButton(master=self.root.frame_1,
                                                 text="חזור",
                                                corner_radius=0,
                                                     command=go_back,
                                                     width=200)
        self.root.button_3.place(relx=0.3, rely=0.85, anchor=tkinter.CENTER)
        try:
            if data_entry != "":
                num = 0
                for data in data_entry:
                    entrys[num].insert(0, str(data))
                    num += 1
        except:
            pass

    def big_title_to_add(self, width, height):
        name = self.the_name_of_the_table.split("_")
        self.root.frame_1.label_1 = customtkinter.CTkLabel(master=self.root.frame_1,
                                                           width=width,
                                                           height=height,
                                                           fg_color=self.color_object,
                                                           corner_radius=0,
                                                           text=name[0],
                                                           text_font=("Halvetica", -16))
        self.root.frame_1.label_1.place(relx=0.5, rely=0.1, anchor=tkinter.CENTER)
    def add_colors(self):
        lis = [[0, 0.03, 20, 1500], [0.05, 0.03, 1500, 20],
               [0.95, 0.03, 1500, 20]]
        for i in range(len(lis)):
            self.root.frame_1.label_1 = customtkinter.CTkLabel(master=self.root.frame_1,
                                                        width=lis[i][3],
                                                  height=lis[i][2],
                                                  corner_radius=0, text="",
                                                               fg_color=self.color_object,
                                                  text_font=("Halvetica", -16))
            self.root.frame_1.label_1.place(relx=lis[i][0], rely=lis[i][1],
                                            anchor=tkinter.CENTER)


class new_wind(Base):
    def __init__(self, root, windows_name, appearance_mode,
                 default_color_theme, box, list_of_buttons, label,
                 the_location, label_2):
        super().__init__(root, windows_name, appearance_mode,
                         default_color_theme, label, the_location)
        self.the_name_of_the_table = "teachers_table"
        self.box = box
        self.root = root
        self.windows_name = windows_name
        self.appearance_mode = appearance_mode
        self.default_color_theme = default_color_theme
        self.list_of_buttons = list_of_buttons
        self.label = label
        self.label_2 = label_2
        self.the_location = the_location


    def Change_the_screen_following_the_table(self, name_for_the_test,
        the_number_of_the_button,the_name_of_the_table):
        self.label_2.configure(text=" ")
        self.box.set(" ")
        self.the_name_of_the_table = the_name_of_the_table
        self.root.label_1.destroy()
        self.root.label_1 = customtkinter.CTkLabel(master=self.root.frame_right,
                                              text=name_for_the_test,
                                              text_font=("Halvetica",
                                                         -25))
        self.root.label_1.place(relx=0.8, rely=0.03)
        for i in range(4):
            if i == the_number_of_the_button:
                self.list_of_buttons[i].configure(fg_color=("gray84", "gray25"))
            else:
                self.list_of_buttons[i].configure(fg_color=("#1f6aa5"))
        list_data = tuple(Base.sever_db(["human", the_name_of_the_table]))
        self.box["values"] = list_data

    def button_pupils(self):
        new_wind.Change_the_screen_following_the_table(self, "תלמידים", 1,
                                                       "pupils_table")

    def button_teacher(self):
        new_wind.Change_the_screen_following_the_table(self, "מורים", 0,
                                                       "teachers_table")

    def button_password(self):
        new_wind.Change_the_screen_following_the_table(self, "סיסמה", 3,
                                                       "password_table")

    def button_Practitioners(self):
        new_wind.Change_the_screen_following_the_table(self, "מתרגלים", 2,
                                               "practitioners_table")

    def remove_item(self):
        typed = self.box.get()
        Base.sever_db(["remove_item", self.the_name_of_the_table,
                               str(typed)])
        self.box["values"] = tuple(Base.sever_db(["human", 
                                                self.the_name_of_the_table]))
        self.box.set("")

    def update_single(self, the_execute):
        if the_execute == "chek":
            data = Base.sever_db(["human", self.the_name_of_the_table])
            global_entry = str(self.label_2.text)
            print(global_entry)
            if global_entry != "":
                for d in data:
                    if d[1] in global_entry:
                        global_entry = d
                        break
                self.root.destroy()
                update_item.update(global_entry, self.the_name_of_the_table)
            else:
                self.label.configure(text="לא בחרת איזה נתונים לעדכן")
                self.root.after(4000, self.clear_label)

        else:
            global_entry = ""
            self.root.destroy()
            update_item.update(global_entry, self.the_name_of_the_table)

    def catching_sound(self, typed):
        data = []
        list_data = Base.sever_db(["human", self.the_name_of_the_table])
        for item in list_data:
            if typed.lower() in item:
                data.append(item)
        self.box["values"] = data

    def excel(self):
        filename = fd.askopenfilename(
            title='Open a file',
            initialdir='/',
        )
        try:
            wb = load_workbook(filename)
            ws = wb.active
            if self.the_name_of_the_table == "teachers_table" or \
                    "practitioners_table" or "pupils_table":
                A = ws["A"]
                B = ws["B"]
                C = ws["C"]
                exel_list = []
                for a in range(len(A)):
                    if A[a].value is None:
                        A[a].value = "ריק"
                    if B[a].value is None:
                        B[a].value = "ריק"
                    if C[a].value is None:
                        C[a].value = "ריק"
            else:
                A = ws["A"]
                B = ws["B"]
                exel_list = []
                for a in range(len(A)):
                    if A[a].value is None:
                        A[a].value = "ריק"
                    if B[a].value is None:
                        B[a].value = "ריק"


            def add_item(item_list):
                for exel in exel_list:
                    bool_i = True
                    for P in item_list:
                        if P[1] == exel[1] and P[0] == exel[0] and P[2] == \
                                exel[2]:
                            bool_i = False
                    if bool_i:
                        item_list.append(exel)
                return item_list

            for i in range(len(A)):
                if self.the_name_of_the_table == "password_table":
                    l = [str(A[i].value), str(B[i].value)]
                else:
                    l = [A[i].value, str(B[i].value), C[i].value]
                exel_list.append(l)
            data_to_add = ["add_exel", exel_list, self.the_name_of_the_table]
            teachers = Base.sever_db(data_to_add)
            teachers = add_item(teachers)
            list_data = tuple(teachers)
            self.box = list_data
            self.label.configure(text="הקובץ שנתנת התווסף בהצלחה")
        except:
            self.label.configure(text="הקובץ שנתנת הוא לא קובץ אקסל")
        self.root.after(4000, self.clear_label)


    def clear_label(self):
        self.label.configure(text="")



    def voice(self):
        def g():
            global recognize
            recognize = False
        recognizer = speech_recognition.Recognizer()
        recognize = True
        with speech_recognition.Microphone() as mic:
            try:
                t_end = time.time() + 15
                while recognize or time.time() > t_end:
                    recognizer.adjust_for_ambient_noise(mic, duration=0.2)
                    audio = recognizer.listen(mic, timeout=0.6)
                    print("audio", audio)
                    if audio != '':
                        self.box.set(str(recognizer.recognize_google(audio,
                                                                 language="he")))
                        self.catching_sound(str(recognizer.recognize_google(audio,
                                                               language="he")))
                        recognize = False
            except:
                self.label = self.create_masages(220, 100, "#2a2d2e",
                                                 "לא הצליח "
                                                                "לזהות "
                                                            "קול",
                                    0.15, 0.3, -16)
                self.root.after(4000, self.clear_label)

    def check(self):
        typed = self.box.get()
        list_data = Base.sever_db(["human", self.the_name_of_the_table])
        if typed == '':
            data = list_data
        else:
            data = []
            for item in list_data:
                for i in item:
                    if typed.lower() in i.lower():
                        data.append(item)
        self.box["values"] = data



